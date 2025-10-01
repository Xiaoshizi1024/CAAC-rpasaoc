# 作者：玩无人机的小狮子
# 版本：1.0.0
# 日期：2025-10-01
# 描述：无人机合格证批量处理系统，支持顺序获取、逐个获取和自动补充错误记录，自动下载并合并图片
print("=== uom运营合格证批量处理系统===")
print("作者：玩无人机的小狮子")

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import os
import traceback
import time
import base64
import re
import subprocess  # 新增：用于调用系统命令清理进程
import platform    # 新增：用于判断操作系统（确保仅Windows生效）
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment

# 基础配置
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_URL = "https://uom.caac.gov.cn/#/uav-sczs-show/"  # 基础URL
EXCEL_FILENAME = "uav_cert_status_correct_range.xlsx"
IMG_XPATHS = [
    '//*[@id="pdfBox"]/div[5]/div/div/div/img',
    '//*[@id="pdfBoxYunxgf"]/div[5]/div/div/div/img'
]
# 核心修改：设置为False，合并后自动删除单张原始图片
KEEP_SINGLE_IMAGES = False  # 仅保留合并后的图片，不保存合并前的单张图片
FAILED_STATUS = {"未找到状态", "获取失败"}  # 完全获取失败的状态集合
# 需要重新处理的状态列表
REPROCESS_STATUSES = {
    "获取失败",
    "未处理(错误: Message: session not)",
    "未处理(错误: HTTPConnectionPool(h)",
    "未处理(错误: Message: Service C:\\)",
    "未处理(错误: Message: Unable to o)",
    "未处理(错误: Message: unknown err)",
    "未处理(错误: [WinError 10054] 远程主)",
    "未处理(错误: Message: invalid ses)"
}

# 需要重新处理的图片合并结果
REPROCESS_MERGE_RESULTS = {"无图片", "失败"}


# 新增：清理Chrome残留进程（针对Windows系统的“Google Chrome for Testing”）
def clean_chrome_processes():
    """清理Windows系统中残留的Chrome测试版进程，避免内存泄漏"""
    # 仅在Windows系统执行（避免跨平台问题）
    if platform.system() != "Windows":
        return
    
    # 需要清理的进程名（精准匹配“Google Chrome for Testing”及关联驱动进程）
    process_names = [
        "Google Chrome for Testing.exe",
        "chromedriver.exe"  # 额外清理可能残留的ChromeDriver进程
    ]
    
    print("\n--- 开始清理Chrome残留进程 ---")
    for proc_name in process_names:
        try:
            # 调用Windows taskkill命令：/F强制结束，/IM按进程名匹配
            subprocess.run(
                ["taskkill", "/F", "/IM", proc_name],
                stdout=subprocess.PIPE,  # 隐藏命令输出
                stderr=subprocess.PIPE,
                shell=True,
                check=True
            )
            print(f"✅ 成功清理进程：{proc_name}")
        except subprocess.CalledProcessError:
            # 进程不存在时会报错，属于正常情况，无需抛出异常
            print(f"ℹ️  未找到残留进程：{proc_name}")
        except Exception as e:
            # 捕获其他异常（如权限问题），避免程序中断
            print(f"⚠️  清理进程{proc_name}失败：{str(e)[:30]}")
    print("--- 进程清理完成 ---\n")


def merge_images(image_paths, output_path, direction='vertical'):
    """合并多张图片（合并后自动删除原始单张图片）"""
    try:
        # 读取所有图片并确保格式一致（转为RGBA避免透明通道问题）
        images = []
        for img_path in image_paths:
            with Image.open(img_path) as img:
                images.append(img.convert("RGBA"))
        
        # 计算合并后图片的尺寸
        if direction == 'vertical':
            max_width = max(img.width for img in images)
            total_height = sum(img.height for img in images)
            merged_img = Image.new('RGBA', (max_width, total_height))
            
            current_y = 0
            for img in images:
                x_offset = (max_width - img.width) // 2
                merged_img.paste(img, (x_offset, current_y), img)
                current_y += img.height
        else:
            total_width = sum(img.width for img in images)
            max_height = max(img.height for img in images)
            merged_img = Image.new('RGBA', (total_width, max_height))
            
            current_x = 0
            for img in images:
                y_offset = (max_height - img.height) // 2
                merged_img.paste(img, (current_x, y_offset), img)
                current_x += img.width
        
        # 保存合并后的图片
        merged_img.save(output_path, format='PNG')
        print(f"图片合并成功！合并后路径：{output_path}")
        
        # 核心逻辑：若不保留单张图片，合并后立即删除原始文件
        if not KEEP_SINGLE_IMAGES:
            for img_path in image_paths:
                if os.path.exists(img_path):  # 确保文件存在再删除
                    os.remove(img_path)
            print(f"已自动删除合并前的单张原始图片（共{len(image_paths)}张）")
        
        return True
    except Exception as e:
        print(f"图片合并失败：{str(e)}")
        print(traceback.format_exc())
        return False


def init_excel(processing_desc):
    """初始化Excel文件（记录处理模式及参数信息）"""
    excel_path = os.path.join(SCRIPT_DIR, EXCEL_FILENAME)
    if os.path.exists(excel_path):
        return excel_path
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "处理结果"
    
    # 记录处理模式及参数（适配顺序/逐个两种模式）
    ws.cell(row=1, column=1, value=processing_desc)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    # 表头
    headers = ["序号", "合格证编号", "完整URL", "状态", "处理结果", "图片合并结果", "耗时(秒)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 30  # 加宽以适应长错误信息
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 10
    
    wb.save(excel_path)
    return excel_path


def update_excel(row_data, row_number=None):
    """更新Excel结果，可指定行号进行更新"""
    try:
        excel_path = os.path.join(SCRIPT_DIR, EXCEL_FILENAME)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["处理结果"]
        
        # 如果指定了行号则更新该行，否则添加新行
        if row_number:
            next_row = row_number
        else:
            next_row = ws.max_row + 1
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
            ws.cell(row=next_row, column=col).alignment = Alignment(
                horizontal='center' if col in [1,4,5,6,7] else 'left', 
                vertical='center'
            )
        
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"更新Excel失败: {str(e)}")
        return False


def extract_cert_status(driver):
    """提取合格证状态"""
    status_xpath = '//*[@id="registerMain"]/div[1]/span[1]'
    
    try:
        status_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, status_xpath))
        )
        
        text_content = status_element.get_attribute("textContent").strip()
        inner_text = status_element.get_attribute("innerText").strip()
        direct_text = status_element.text.strip()
        
        status_candidates = [text_content, inner_text, direct_text]
        for text in status_candidates:
            cleaned_text = re.sub(r'\s+', '', text)
            if cleaned_text == "已启用":
                return "已启用"
            elif cleaned_text == "已撤销":
                return "已撤销"
        
        return f"未知({status_candidates[0][:10]})"
    except Exception as e:
        try:
            page_source = driver.page_source
            if re.search(r'已启用', page_source):
                return "已启用(源码)"
            elif re.search(r'已撤销', page_source):
                return "已撤销(源码)"
            else:
                return "未找到状态"
        except:
            return "获取失败"


def process_single_cert(cert_number, wait_time=10):
    """处理单个合格证（单张图片临时下载，合并后自动删除）"""
    start_time = time.time()
    result = {
        "success": False,
        "status": "未处理",
        "url": f"{BASE_URL}{cert_number}",
        "merge_result": "未合并"
    }
    
    driver = None
    try:
        # 初始化浏览器（无头模式）
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--disable-blink-features=AutomationControlled")
        driver = webdriver.Chrome(options=options)
        
        # 访问链接并提取状态
        print(f"访问: {result['url']}")
        driver.get(result["url"])
        time.sleep(wait_time)
        result["status"] = extract_cert_status(driver)
        print(f"状态提取结果: {result['status']}")
        
        # 状态完全获取失败则跳过图片操作
        if result["status"] in FAILED_STATUS:
            print(f"⚠️  状态属于完全获取失败，跳过图片下载与合并")
            result["merge_result"] = "跳过（状态失败）"
            result["success"] = False
            return result
        
        # 临时下载单张图片（合并后会删除）
        save_dir = os.path.join(SCRIPT_DIR, "uav_cert_images")
        os.makedirs(save_dir, exist_ok=True)
        downloaded_images = []
        
        print("开始定位并临时下载图片（合并后自动删除）...")
        for idx, xpath in enumerate(IMG_XPATHS, 1):
            print(f"\n--- 处理第{idx}张临时图片 (XPath: {xpath}) ---")
            try:
                img_element = WebDriverWait(driver, 60).until(
                    EC.visibility_of_element_located((By.XPATH, xpath))
                )
                print(f"第{idx}张图片元素已找到")
                
                img_src = img_element.get_attribute("src")
                print(f"第{idx}张图片源: {img_src[:50]}...")
                
                if not img_src:
                    print(f"第{idx}张图片未获取到源数据，跳过")
                    continue
                
                # 临时保存单张图片（后续会删除）
                temp_img_path = os.path.join(save_dir, f"{cert_number}_temp_img_{idx}.png")
                if img_src.startswith('data:image/'):
                    print("检测到Base64格式，开始解码（临时保存）...")
                    header, base64_data = img_src.split(',', 1)
                    image_data = base64.b64decode(base64_data)
                    with open(temp_img_path, "wb") as f:
                        f.write(image_data)
                else:
                    print("检测到URL格式，开始下载（临时保存）...")
                    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
                    response = requests.get(img_src, headers=headers, timeout=30)
                    response.raise_for_status()
                    with Image.open(BytesIO(response.content)) as img:
                        img.save(temp_img_path, format='PNG')
                
                print(f"第{idx}张临时图片已保存: {temp_img_path}")
                downloaded_images.append(temp_img_path)
            
            except Exception as e:
                print(f"第{idx}张图片处理失败: {str(e)}")
                print(traceback.format_exc())
        
        # 图片合并（合并后自动删除临时文件）
        if downloaded_images:
            print("\n=== 开始图片合并（合并后删除临时文件）===")
            merge_path = os.path.join(save_dir, f"{cert_number}_merged.png")
            if merge_images(downloaded_images, merge_path, direction='vertical'):
                result["merge_result"] = "成功（仅保留合并图）"
            else:
                result["merge_result"] = "失败"
        else:
            print("\n无有效图片可合并")
            result["merge_result"] = "无图片"
        
        # 状态正常（含未知）视为处理成功
        result["success"] = True
            
    except Exception as e:
        result["status"] += f"(错误: {str(e)[:20]})"
        result["merge_result"] = "中断"
        result["success"] = False
    finally:
        # 1. 优先关闭浏览器实例
        if driver:
            driver.quit()
            print(f"已关闭当前Chrome浏览器实例")
        # 2. 新增：清理可能残留的Chrome进程（关键修复内存泄漏）
        clean_chrome_processes()
        # 3. 计算耗时并返回结果
        result["time"] = round(time.time() - start_time, 2)
        return result


def generate_cert_numbers(start, end, month_prefix):
    """生成带自定义年月份前缀的合格证编号（顺序模式用）"""
    numbers = []
    for num in range(start, end + 1):
        numbers.append(f"BZSQ914{month_prefix}{num:03d}")
    return numbers


def get_valid_month_prefix():
    """获取并验证用户输入的年月份前缀（4位数字，顺序模式用）"""
    while True:
        prefix = input("\n请输入年月份前缀（例如BZSQ9142401001中的'2401'）：").strip()
        if re.match(r'^\d{4}$', prefix):
            return prefix
        print("输入格式错误！请输入4位数字（例如2401）")


def get_valid_number(prompt, min_value):
    """获取并验证用户输入的数字（顺序模式用）"""
    while True:
        try:
            num = int(input(prompt).strip())
            if num >= min_value:
                return num
            print(f"输入错误！请输入不小于{min_value}的整数")
        except ValueError:
            print("输入错误！请输入有效的整数")


def get_reprocess_cert_numbers():
    """从Excel文件中读取需要重新处理的合格证编号"""
    excel_path = os.path.join(SCRIPT_DIR, EXCEL_FILENAME)
    if not os.path.exists(excel_path):
        print(f"错误：未找到Excel文件 {excel_path}")
        return []
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["处理结果"]
        
        # 存储需要重新处理的编号及其行号
        reprocess_items = []
        
        # 从第3行开始读取数据（前两行是标题和表头）
        for row in range(3, ws.max_row + 1):
            cert_number = ws.cell(row=row, column=2).value
            status = ws.cell(row=row, column=4).value
            merge_result = ws.cell(row=row, column=6).value
            
            # 检查是否需要重新处理
            need_reprocess = False
            if status in REPROCESS_STATUSES:
                need_reprocess = True
            elif merge_result in REPROCESS_MERGE_RESULTS:
                need_reprocess = True
                
            if need_reprocess and cert_number:
                reprocess_items.append((cert_number, row))
        
        print(f"发现 {len(reprocess_items)} 条需要重新处理的记录")
        return reprocess_items
    except Exception as e:
        print(f"读取Excel文件失败: {str(e)}")
        return []


def main():
    # 处理模式选择
    print("\n请选择处理模式：")
    print("1. 顺序获取（按编号范围批量生成，如2401001-2401010）")
    print("2. 逐个获取（手动输入错误编号补充，用英文逗号分隔）")
    print("3. 自动获取（自动补充xlsx表内错误）")
    
    # 验证模式选择
    while True:
        mode_choice = input("请输入1、2或3选择模式：").strip()
        if mode_choice in ["1", "2", "3"]:
            break
        print("输入无效！请仅输入1、2或3")

    # 按模式获取合格证编号列表及处理描述
    if mode_choice == "1":
        # 原有顺序模式逻辑
        month_prefix = get_valid_month_prefix()
        start_num = get_valid_number("\n请输入起始数字（例如1）：", 1)
        end_num = get_valid_number(f"请输入结束数字（不小于{start_num}）：", start_num)
        cert_numbers = generate_cert_numbers(start_num, end_num, month_prefix)
        processing_desc = (f"模式：顺序获取 | 年月份前缀：{month_prefix} | 编号范围：{start_num}-{end_num} "
                          f"| 图片策略：仅保留合并后图片")
        # 对于顺序模式，存储的是编号列表
        process_items = [(num, None) for num in cert_numbers]
        
    elif mode_choice == "2":
        # 逐个模式逻辑：解析逗号分隔的编号
        while True:
            input_str = input("\n请输入多个合格证编号（用英文逗号分隔，如BZSQ9142401001,BZSQ9142401003）：").strip()
            # 分割+去空格+去重（保留输入顺序）
            cert_list = [item.strip() for item in input_str.split(",") if item.strip()]
            cert_numbers = list(dict.fromkeys(cert_list))  # 去重且保持顺序
            if cert_numbers:
                break
            print("输入无效！请至少输入1个编号，且编号间用英文逗号分隔")
        processing_desc = (f"模式：逐个获取 | 编号数量：{len(cert_numbers)} | 编号列表：{','.join(cert_numbers)} "
                          f"| 图片策略：仅保留合并后图片")
        # 对于逐个模式，存储的是编号列表
        process_items = [(num, None) for num in cert_numbers]
        
    else:  # mode_choice == "3"
        # 自动获取模式：从Excel读取需要重新处理的编号
        process_items = get_reprocess_cert_numbers()
        if not process_items:
            print("没有需要重新处理的记录，程序退出")
            return
            
        cert_numbers = [item[0] for item in process_items]
        processing_desc = (f"模式：自动获取 | 重新处理数量：{len(cert_numbers)} | "
                          f"图片策略：仅保留合并后图片")

    # 显示参数确认
    print(f"\n=== 参数确认 ===")
    print(processing_desc)

    # 初始化Excel（传入模式描述）
    excel_path = init_excel(processing_desc)
    total = len(process_items)
    success_count = 0
    fail_count = 0

    print(f"\n开始处理: 共{total}个链接")
    if total > 0:
        print(f"链接示例: {BASE_URL}{process_items[0][0]}")

    # 逐个处理编号
    for i, (cert_number, row_number) in enumerate(process_items, 1):
        print(f"\n===== 处理 {i}/{total}: {cert_number} =====")
        result = process_single_cert(cert_number, wait_time=10)

        # 更新统计
        if result["success"]:
            success_count += 1
        else:
            fail_count += 1

        # 更新Excel，自动模式下更新原有行，其他模式添加新行
        update_excel([
            i if row_number is None else row_number - 2,  # 保持原有序号
            cert_number,
            result["url"],
            result["status"],
            "成功" if result["success"] else "失败",
            result["merge_result"],
            result["time"]
        ], row_number)

        # 进度提示
        print(f"状态: {result['status']} | 合并: {result['merge_result']} | 耗时: {result['time']}秒 | 结果: {'成功' if result['success'] else '失败'}")
        print(f"进度: 成功{success_count}, 失败{fail_count}, 总计{success_count+fail_count}/{total}")

    # 新增：程序结束前再次清理进程（双重保障）
    print("\n--- 程序即将结束，最终清理进程 ---")
    clean_chrome_processes()

    # 完成总结
    print("\n===== 处理完成 =====")
    print(f"总数量: {total}")
    print(f"成功: {success_count} ({round(success_count/total*100, 2)}%)")
    print(f"失败: {fail_count} ({round(fail_count/total*100, 2)}%)")
    print(f"结果文件: {excel_path}")
    print(f"图片目录: {os.path.join(SCRIPT_DIR, 'uav_cert_images')}（仅含合并后的图片）")

if __name__ == "__main__":
    main()
